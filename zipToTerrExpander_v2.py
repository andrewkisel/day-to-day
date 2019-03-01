#! python3
# This app expands zip-code ranges in zip-to-terr files,
# removes unneeded columns and performs lookup of city based on zip-code.

import os
import sys
import openpyxl
import logging
from zipCodeDB import CA_ZIPS, US_ZIPS, ZIP_CITY
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import re

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.DEBUG)

# DISCLAIMER:
# 1. In order to use the script, place all of the zip-to-terr files in one folder. Each file must contain either
# multiple sheets along with the 'Geo Assignment' one, but it should be active (when you open the spreadsheet - it
# should open first, or single 'Geo Assignment' spreadsheet.
# 2. Names of the columns with zip-codes should end with 'Low' and 'High' respectively.
# 3. Names of the files should follow this pattern: 2018_NA_Applied_LW_Territory_Rules.xlsx (example)

disclaimer = '''
****************************************************************************************************************

* DISCLAIMER:
* In order to use the script, place all of the zip-to-terr files in one folder. Each file must contain either
multiple sheets along with the 'Geo Assignment' one, but it should be active (when you open the spreadsheet - it
should open first), or single 'Geo Assignment' spreadsheet.
* Names of the columns with zip-codes should end with 'Low' and 'High' respectively.
* Names of the files should follow this pattern: 2018_NA_Applied_LW_Territory_Rules.xlsx (example)

****************************************************************************************************************
'''
file_name_rex = re.compile(r'(^\d+_)(\w+_)(\w+|\w+ \(\w+\)|\w+ \w+|\w+\(\w+\))(_Territory_)(Rules)(\.xlsx$)',
                           re.IGNORECASE)


# Expand the zip-code ranges.
def zip_expander(path):
    # Change working directory to file path and open the spreadsheet.
    os.chdir(os.path.dirname(path))
    file_name = os.path.basename(path)
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    # Prepare result file + worksheet.
    res_wb = openpyxl.Workbook()
    res_ws = res_wb.active
    # Identify the rep type
    assert file_name_rex.search(file_name) is not None, 'Cannot locate the rep type in file name. ' \
                                                        'Names of the files should follow this pattern: ' \
                                                        'any text_Specialist type_Territory_Rules.xlsx'
    rep_type = file_name_rex.search(file_name).group(3)
    # Result file row count
    res_row = 1

    # Loop through rows in spreadsheet.
    for row in range(ws.min_row, ws.max_row + 1):
        logging.debug('Processing row: %d' % row)
        # First row is header. Copy it along with making changes.
        if row == 1:
            for head in ws.iter_cols(min_row=1, max_row=1, max_col=ws.max_column):
                for cell in head:
                    # Identify the zip-code ranges in data set.
                    if re.compile(r'.*low$', re.IGNORECASE).search(cell.value):
                        start_zip_col = cell.column
                        # Rename the zip-code column.
                        res_ws[cell.coordinate] = 'Align Star Zip-Code'
                    elif re.compile(r'.*high$', re.IGNORECASE).search(cell.value):
                        end_zip_col = cell.column
                    elif re.compile(r'^Territory.*Code$', re.IGNORECASE).search(cell.value):
                        res_ws[cell.coordinate] = rep_type + ' Rep Territory ID'
                    else:
                        res_ws[cell.coordinate] = cell.value
            res_row += 1
            continue
        # On each iteration store low and high zip-codes here.
        assert start_zip_col and end_zip_col is not None, 'Could not find the columns of zip-code ranges. ' \
                                                          'They should be named as <sometext_Low>' \
                                                          ' and <sometext_High> respectively.'
        start_code = str(ws[start_zip_col + str(row)].value)
        end_code = str(ws[end_zip_col + str(row)].value)
        # In case some US zip-codes lost leading zeros - add them as necessary.
        if len(start_code) < 5 and len(end_code) < 5 and start_code.isdigit():
            start_code = (5 - len(start_code)) * '0' + start_code
            end_code = (5 - len(end_code)) * '0' + end_code
        logging.debug('Start code is: %s' % start_code)
        logging.debug('End code is: %s' % end_code)
        # Identify if low/high zip-code is US or CA one.
        if start_code in US_ZIPS and end_code in US_ZIPS:
            # Make sure if high zip-code is bigger than low.
            assert US_ZIPS.index(end_code) >= US_ZIPS.index(
                start_code), 'Check your ranges! Postal code high: %s is less then postal code low: %s' % (
                end_code, start_code)
            # Loop through list of all US zip-codes to expand the ranges. Write to file.
            for i in range(US_ZIPS.index(start_code), len(US_ZIPS)):
                logging.debug('List index is: %d' % i)
                logging.debug('Result file row is: %d' % res_row)
                res_ws[start_zip_col + str(res_row)] = US_ZIPS[i]
                # Copy other columns on each iteration. Write to file.
                for col in range(ws.min_column, ws.max_column + 1):
                    col = get_column_letter(col)
                    # Except start/end zip columns.
                    if col == start_zip_col or col == end_zip_col:
                        continue
                    res_ws[col + str(res_row)] = ws[col + str(row)].value
                # Once reached end-code in DB - write it to file, break out of the loop.
                if US_ZIPS[i] == end_code:
                    res_ws[start_zip_col + str(res_row)] = end_code
                    res_row += 1
                    break
                res_row += 1
        # Logic here is absolutely the same as with US zips.
        elif start_code in CA_ZIPS and end_code in CA_ZIPS:
            assert CA_ZIPS.index(end_code) >= CA_ZIPS.index(
                start_code), 'Check your ranges! Postal code high: %s is less then postal code low: %s' % (
                end_code, start_code)
            for i in range(CA_ZIPS.index(start_code), len(CA_ZIPS)):
                logging.debug('List index is: %d' % i)
                logging.debug('Result file row is: %d' % res_row)
                res_ws[start_zip_col + str(res_row)] = CA_ZIPS[i]
                for col in range(ws.min_column, ws.max_column + 1):
                    col = get_column_letter(col)
                    if col == start_zip_col or col == end_zip_col:
                        continue
                    res_ws[col + str(res_row)] = ws[col + str(row)].value
                if CA_ZIPS[i] == end_code:
                    res_ws[start_zip_col + str(res_row)] = end_code
                    res_row += 1
                    break
                res_row += 1
        # In case some of the zip-codes is not in the DB - show warning and continue to next iteration.
        else:
            print(
                '* WARNING: Unable to process row: %s. Low zip-code: %s, or high zip-code: %s is not in the database' %
                    (row, start_code, end_code))
            continue
    # Delete zip-code high column from result file.
    res_ws.delete_cols(column_index_from_string(end_zip_col), amount=1)
    res_ws.title = rep_type
    # Return result file
    return city_filler(res_wb)


# Identify the irrelevant columns and remove them.
def column_remover(workbook):
    print('* Removing irrelevant columns...')
    ws = workbook.active
    for index in range(1, ws.max_column + 1):
        for i in range(1, ws.max_column + 1):
            if not re.compile(r'(.*Rep Territory ID$)|(City)|(.*State$)|(Align Star Zip-Code)', re.IGNORECASE).search(
                    str(ws[get_column_letter(i) + '1'].value)):
                ws.delete_cols(i, amount=1)
                break
    return workbook


# Fill in the missing city column
def city_filler(workbook):
    print('* Adding city records based on zip-codes...')
    ws = workbook.active
    raw_zip_list = []
    zip_col_id = None
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, max_row=1):
        for cell in col:
            if ws[cell.coordinate].value == 'Align Star Zip-Code':
                zip_col_id = column_index_from_string(cell.column)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=zip_col_id, max_col=zip_col_id):
        for cell in row:
            cur_code = str(cell.value)
            raw_zip_list.append(cur_code)
    ws.insert_cols(zip_col_id, amount=1)
    ws[get_column_letter(zip_col_id) + '1'] = 'City'
    city_list = bulk_zip_lookup(raw_zip_list)
    for i in range(len(city_list)):
        if city_list[i] == 'Not Found':
            print('* WARNING: Item in row: %s is not in the database' % str(i + 2))
        ws[get_column_letter(zip_col_id) + str(i + 2)] = city_list[i]
    return column_remover(workbook)


# Lookup city based on zip-codes
def bulk_zip_lookup(raw_data):
    res_data = []
    # loop through incoming list.
    for item in raw_data:
        if ZIP_CITY.get(item) is not None:
            # Where zip-codes lost zeros - add them.
            if len(item) < 5 and item.isdigit():
                item = (5 - len(item)) * '0' + item
            # Save to resulting list.
            res_data.append(ZIP_CITY.get(item).title())
        # Otherwise indicate that item has not been found.
        else:
            res_data.append('Not Found')
    return res_data


assert len(sys.argv) == 2, 'Please check the file path provided.'
print(disclaimer)
file_path = sys.argv[1]
raw_files = os.listdir(file_path)
for file in raw_files:
    if not file_name_rex.search(file):
        print('* WARNING: File "%s" does not follow the file name pattern. Skipping it...' % file)
        continue
    print('* Processing file %s...' % file)
    zip_expander(os.path.join(file_path, file)).save('Expanded_' + file)
print('* Process complete.')
