#! python3
# Splits the file into 500 chunks, makes DT scenarios and job file.

import csv
import openpyxl

dt_job_text = r'''Scenarios_XML_Here'''

path = 'C:\\Users\\User\\Desktop\\merge_zeros\\merge2.xlsx'
print('Loading workbook...')
raw_wb = openpyxl.load_workbook(path)
raw_ws = raw_wb.active
top_row = ['Master ID', 'Dupe ID']
batch_file = open('C:\\Users\\User\\Desktop\\merge_zeros\\run_1\\batch.djb', 'w')
batch_file.writelines(r'''JobBuilder 3.0
RunDemandTools
Set by DemandTools
"C:\Users\User\Desktop\merge_zeros\run_1\1_500.STDxml"
10000''')
for it in range(1, raw_ws.max_row):
    if it % 500 == 0:
        print('Processing chunk: %d ... %d' % ((it - 499), it))
        csv_file = open('C:\\Users\\User\\Desktop\\merge_zeros\\run_1\\%d_%d.csv' % ((it - 499), it), 'w',
                        newline='')
        job_file = open('C:\\Users\\User\\Desktop\\merge_zeros\\run_1\\%d_%d.STDxml' % ((it - 499), it), 'w')
        csv_writer = csv.writer(csv_file)
        csv_writer.writerow(top_row)
        for row in raw_ws.iter_rows(min_row=it - 499, max_row=it, max_col=2):
            rowData = []
            for cell in row:
                rowData.append(cell.value)
            csv_writer.writerow(rowData)
        job_file.writelines(dt_job_text.replace('place_for_file_name', str(it - 499) + '_' + str(it)))
        if it / 500 != 1:
            batch_file.writelines('\n\n\n' + r'''X
DemandTools as last logged in
RunDemandTools
Set by DemandTools
"C:\Users\User\Desktop\merge_zeros\run_1\%d_%d.STDxml"
10000''' % (it - 499, it))
        job_file.close()
        csv_file.close()
batch_file.close()
